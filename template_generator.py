# template_generator.py
import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]

FILL_HEADER      = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER_MID  = PatternFill("solid", fgColor="2E75B6")
FILL_HEADER_LIGHT= PatternFill("solid", fgColor="BDD7EE")
FILL_EXAMPLE     = PatternFill("solid", fgColor="EBF3FB")
FILL_NOTE        = PatternFill("solid", fgColor="FFF2CC")
FILL_URGENT      = PatternFill("solid", fgColor="FFE0E0")
FILL_LOCKED      = PatternFill("solid", fgColor="F2F2F2")

FONT_WHITE       = Font(bold=True, color="FFFFFF", size=11)
FONT_WHITE_SM    = Font(bold=True, color="FFFFFF", size=10)
FONT_BOLD        = Font(bold=True, size=10)
FONT_ITALIC      = Font(italic=True, color="7F7F7F", size=9)
FONT_NORMAL      = Font(size=10)

ALIGN_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
MEDIUM_BORDER = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)


def _cell(ws, row, col, value="", fill=None, font=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border
    return c


def _add_petunjuk_sheet(wb):
    ws = wb.create_sheet("📋 Petunjuk")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 65

    _cell(ws, 1, 1, "📋 PETUNJUK PENGGUNAAN TEMPLATE",
          font=Font(bold=True, size=13))
    ws.merge_cells("A1:B1")

    rows = [
        ("FILE", "KETERANGAN"),
        ("Master Mixer",     "Daftar semua mixer: nama, kapasitas (kg/batch), batch per shift, grup cleaning."),
        ("Master Produk",    "Daftar produk: kode, nama, kg/CS, resting days, grup cleaning, mixer kompatibel."),
        ("Filling Plan",     "Rencana filling: produk (baris) vs tanggal+shift (kolom). Isi Target CS di sel perpotongan."),
        ("", ""),
        ("KOLOM", "KETERANGAN"),
        ("Kode_Produk",      "Kode unik produk — harus cocok dengan Master Produk."),
        ("Nama_Produk",      "Nama produk (otomatis, tidak perlu diisi ulang — hanya referensi)."),
        ("Urgent",           "Isi 'Urgent' untuk prioritas tinggi, kosong / 'Normal' untuk biasa."),
        ("Senin S1 / S2 / S3", "Isi jumlah Target CS di kolom tanggal+shift yang sesuai. Kosongkan jika tidak ada filling."),
        ("", ""),
        ("CATATAN", ""),
        ("Format tanggal",   "Header kolom sudah otomatis. Cukup isi angka CS di sel yang sesuai."),
        ("Baris contoh",     "Hapus baris contoh sebelum upload, atau biarkan (baris kosong diabaikan)."),
        ("Kolom Nama_Produk","Tidak perlu diisi, hanya referensi visual."),
    ]

    for ri, (a, b) in enumerate(rows, 3):
        ca = ws.cell(row=ri, column=1, value=a)
        cb = ws.cell(row=ri, column=2, value=b)
        if a in ("FILE", "KOLOM", "CATATAN"):
            ca.font = FONT_BOLD
            cb.font = FONT_BOLD
            ca.fill = PatternFill("solid", fgColor="D9E1F2")
            cb.fill = PatternFill("solid", fgColor="D9E1F2")
        ca.alignment = ALIGN_LEFT
        cb.alignment = ALIGN_LEFT

    ws.row_dimensions[1].height = 22


def generate_template_master_mixer():
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Mixer"

    cols = ["Mixer", "Kapasitas_kg", "Batch_per_Shift", "Grup_Cleaning"]
    for ci, h in enumerate(cols, 1):
        _cell(ws, 1, ci, h, fill=FILL_HEADER, font=FONT_WHITE,
              align=ALIGN_CENTER, border=THIN_BORDER)

    examples = [
        ["M1", 500, 2, "A"],
        ["M2", 500, 2, "A"],
        ["M3", 300, 3, "B"],
        ["M4", 400, 2, "B"],
    ]
    for ri, ex in enumerate(examples, 2):
        for ci, val in enumerate(ex, 1):
            _cell(ws, ri, ci, val, fill=FILL_EXAMPLE,
                  font=FONT_NORMAL, align=ALIGN_LEFT, border=THIN_BORDER)

    notes = ["Nama mixer unik", "kg/batch", "batch/shift", "Grup cleaning (A/B/C/...)"]
    for ci, val in enumerate(notes, 1):
        _cell(ws, len(examples)+2, ci, val, fill=FILL_NOTE,
              font=FONT_ITALIC, align=ALIGN_LEFT, border=THIN_BORDER)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.row_dimensions[1].height = 22

    _add_petunjuk_sheet(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_template_master_produk():
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Produk"

    cols = [
        "Kode_Produk", "Nama_Produk", "Kode_MC_Liquid",
        "Kg_per_CS", "Resting_Days", "Grup_Cleaning", "Mixer_Kompatibel"
    ]
    for ci, h in enumerate(cols, 1):
        _cell(ws, 1, ci, h, fill=FILL_HEADER, font=FONT_WHITE,
              align=ALIGN_CENTER, border=THIN_BORDER)

    examples = [
        ["PRD001", "Produk A 1L",    "MC-001", 12.5, 1, "A", "M1,M2"],
        ["PRD002", "Produk B 500ml", "MC-002", 6.0,  0, "A", "M1,M2,M3"],
        ["PRD003", "Produk C 2L",    "MC-003", 24.0, 2, "B", "M3,M4"],
        ["PRD004", "Produk D 250ml", "MC-004", 3.0,  0, "B", "M4"],
    ]
    for ri, ex in enumerate(examples, 2):
        for ci, val in enumerate(ex, 1):
            _cell(ws, ri, ci, val, fill=FILL_EXAMPLE,
                  font=FONT_NORMAL, align=ALIGN_LEFT, border=THIN_BORDER)

    notes = [
        "Kode unik", "Nama produk", "Kode MC (opsional)",
        "kg per CS", "Hari resting (0=hari H)", "Grup cleaning", "Pisah koma"
    ]
    for ci, val in enumerate(notes, 1):
        _cell(ws, len(examples)+2, ci, val, fill=FILL_NOTE,
              font=FONT_ITALIC, align=ALIGN_LEFT, border=THIN_BORDER)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 22
    ws.row_dimensions[1].height = 22

    _add_petunjuk_sheet(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_template_filling_plan(n_days=14, start_date=None):
    """
    Template Filling Plan dengan layout kalender.
    Baris  = produk (Kode_Produk | Nama_Produk | Urgent)
    Kolom  = setiap (tanggal, shift) → user isi Target CS

    Setelah diupload, app.py perlu melt/unpivot dulu sebelum masuk scheduler.
    """
    if start_date is None:
        # Mulai dari Senin minggu ini
        today = datetime.today()
        start_date = today - timedelta(days=today.weekday())

    wb = Workbook()
    ws = wb.active
    ws.title = "Filling Plan"

    # ── Bangun daftar (date, shift) ───────────────────────────────────────────
    date_shift_cols = []
    for d in range(n_days):
        dt = start_date + timedelta(days=d)
        for s in [1, 2, 3]:
            date_shift_cols.append((dt, s))

    FIXED = 3   # kolom tetap: Kode_Produk, Nama_Produk, Urgent

    # ── Row 1: header tanggal (merge 3 kolom per tanggal) ────────────────────
    col_cursor = FIXED + 1
    current_date = None
    for (dt, s) in date_shift_cols:
        if dt != current_date:
            current_date = dt
            day_name = DAYS_ID[dt.weekday()]
            label    = f"{day_name}\n{dt.strftime('%d/%m/%Y')}"
            c = ws.cell(row=1, column=col_cursor, value=label)
            c.fill      = FILL_HEADER
            c.font      = FONT_WHITE
            c.alignment = ALIGN_CENTER
            c.border    = THIN_BORDER
            ws.merge_cells(
                start_row=1, start_column=col_cursor,
                end_row=1,   end_column=col_cursor + 2
            )
        col_cursor += 1

    # ── Row 2: header kolom tetap + shift ────────────────────────────────────
    fixed_headers = ["Kode_Produk", "Nama_Produk", "Urgent"]
    for ci, h in enumerate(fixed_headers, 1):
        _cell(ws, 2, ci, h, fill=FILL_HEADER, font=FONT_WHITE,
              align=ALIGN_CENTER, border=THIN_BORDER)
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

    for ci, (dt, s) in enumerate(date_shift_cols, FIXED + 1):
        _cell(ws, 2, ci, f"S{s}", fill=FILL_HEADER_LIGHT, font=FONT_BOLD,
              align=ALIGN_CENTER, border=THIN_BORDER)

    # ── Row 3: baris catatan ─────────────────────────────────────────────────
    note_row = 3
    _cell(ws, note_row, 1, "← Kode harus ada di Master Produk",
          fill=FILL_NOTE, font=FONT_ITALIC, align=ALIGN_LEFT, border=THIN_BORDER)
    _cell(ws, note_row, 2, "← Nama (opsional, hanya referensi)",
          fill=FILL_NOTE, font=FONT_ITALIC, align=ALIGN_LEFT, border=THIN_BORDER)
    _cell(ws, note_row, 3, "'Urgent' / kosong",
          fill=FILL_NOTE, font=FONT_ITALIC, align=ALIGN_LEFT, border=THIN_BORDER)
    for ci in range(FIXED + 1, FIXED + len(date_shift_cols) + 1):
        _cell(ws, note_row, ci, "← isi CS",
              fill=FILL_NOTE, font=FONT_ITALIC, align=ALIGN_CENTER, border=THIN_BORDER)

    # ── Row 4-8: contoh data ─────────────────────────────────────────────────
    example_products = [
        ("PRD001", "Produk A 1L",    "Normal"),
        ("PRD002", "Produk B 500ml", "Urgent"),
        ("PRD003", "Produk C 2L",    "Normal"),
        ("PRD004", "Produk D 250ml", "Normal"),
        ("PRD001", "Produk A 1L",    "Urgent"),
    ]
    example_cs = [100, 200, 150, 80, 120]

    for ri, ((kode, nama, urgent), cs) in enumerate(
        zip(example_products, example_cs), note_row + 1
    ):
        fill = FILL_URGENT if urgent == "Urgent" else FILL_EXAMPLE
        _cell(ws, ri, 1, kode,   fill=fill, font=FONT_NORMAL, align=ALIGN_LEFT,  border=THIN_BORDER)
        _cell(ws, ri, 2, nama,   fill=fill, font=FONT_NORMAL, align=ALIGN_LEFT,  border=THIN_BORDER)
        _cell(ws, ri, 3, urgent, fill=fill, font=FONT_NORMAL, align=ALIGN_CENTER, border=THIN_BORDER)
        # Isi satu sel CS di shift/hari pertama sebagai contoh
        col_example = FIXED + 1 + ri - (note_row + 1)
        col_example = min(col_example, FIXED + len(date_shift_cols))
        _cell(ws, ri, col_example, cs,
              fill=PatternFill("solid", fgColor="D9EAD3"),
              font=Font(bold=True, size=10),
              align=ALIGN_CENTER, border=THIN_BORDER)
        # Sisa sel kosong
        for ci in range(FIXED + 1, FIXED + len(date_shift_cols) + 1):
            if ci != col_example:
                _cell(ws, ri, ci, "", fill=FILL_LOCKED,
                      font=FONT_NORMAL, align=ALIGN_CENTER, border=THIN_BORDER)

    # ── Row kosong untuk input user (30 baris) ────────────────────────────────
    data_start = note_row + len(example_products) + 1
    for ri in range(data_start, data_start + 30):
        for ci in range(1, FIXED + 1):
            _cell(ws, ri, ci, "", fill=None, font=FONT_NORMAL,
                  align=ALIGN_LEFT, border=THIN_BORDER)
        for ci in range(FIXED + 1, FIXED + len(date_shift_cols) + 1):
            _cell(ws, ri, ci, "", fill=None, font=FONT_NORMAL,
                  align=ALIGN_CENTER, border=THIN_BORDER)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    for ci in range(FIXED + 1, FIXED + len(date_shift_cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 7

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[note_row].height = 16
    ws.freeze_panes = f"{get_column_letter(FIXED + 1)}3"

    _add_petunjuk_sheet(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
