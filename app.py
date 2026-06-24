import streamlit as st
import pandas as pd
import io
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from scheduler import generate_mixing_schedule
from pivot import build_pivot, pivot_to_excel

st.set_page_config(page_title="Mixing Scheduler", page_icon="🧪", layout="wide")
st.title("🧪 Mixing Schedule Planner")

DAYS_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]

# ─── Session State ─────────────────────────────────────────────────────────
if "master_mixer" not in st.session_state:
    st.session_state.master_mixer = pd.DataFrame(
        columns=["Mixer", "Kapasitas_kg", "Grup_Cleaning"])
if "master_produk" not in st.session_state:
    st.session_state.master_produk = pd.DataFrame(
        columns=["Kode_Produk", "Nama_Produk", "Grup_Cleaning",
                 "Kg_per_CS", "Resting_Days", "Mixer_Kompatibel"])
if "filling_plan" not in st.session_state:
    st.session_state.filling_plan = pd.DataFrame()

tab1, tab2, tab3 = st.tabs(
    ["⚙️ Master Data", "📋 Input Planning", "📅 Jadwal Mixing"])

# ═══════════════════════════════════════════════════════════════════════════
# TAB 1 — MASTER DATA
# ═══════════════════════════════════════════════════════════════════════════
with tab1:
    st.header("Master Data")
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("🔧 Master Mixer")
        tmpl_mixer = pd.DataFrame({
            "Mixer":           ["Mixer A", "Mixer B", "Mixer C"],
            "Kapasitas_kg":    [500, 800, 300],
            "Batch_per_Shift": [2, 3, 2],
            "Grup_Cleaning":   ["Grup 1", "Grup 1", "Grup 2"]
        })
        buf = io.BytesIO()
        tmpl_mixer.to_excel(buf, index=False)
        st.download_button(
            "📥 Download Template Mixer", buf.getvalue(),
            "template_master_mixer.xlsx", use_container_width=True)

        up_mixer = st.file_uploader(
            "Upload Master Mixer", type=["xlsx", "csv"], key="mixer_upload")
        if up_mixer:
            df = (pd.read_excel(up_mixer) if up_mixer.name.endswith("xlsx")
                  else pd.read_csv(up_mixer))
            req = {"Mixer", "Kapasitas_kg", "Batch_per_Shift", "Grup_Cleaning"}
            if req.issubset(df.columns):
                st.session_state.master_mixer = df
                st.success(f"✅ {len(df)} mixer berhasil diupload!")
            else:
                st.error(f"❌ Kolom harus: {req}")

        if not st.session_state.master_mixer.empty:
            st.dataframe(st.session_state.master_mixer,
                         use_container_width=True, hide_index=True)

    with col2:
        st.subheader("📦 Master Produk")
        st.caption(
            "**Resting_Days**: `2` = perlu didiamkan 2 hari, `0` = tidak. "
            "**Mixer_Kompatibel**: pisah koma.")
        tmpl_produk = pd.DataFrame({
            "Kode_Produk":      ["P001", "P002", "P003"],
            "Nama_Produk":      ["Produk Alpha", "Produk Beta", "Produk Gamma"],
            "Kode_MC_Liquid":   ["ML001", "ML002", "ML001"],
            "Grup_Cleaning":    ["Grup 1", "Grup 1", "Grup 2"],
            "Kg_per_CS":        [12.5, 8.0, 10.0],
            "Resting_Days":     [0, 2, 0],
            "Mixer_Kompatibel": ["Mixer A, Mixer B", "Mixer B", "Mixer C"]
        })
        buf2 = io.BytesIO()
        tmpl_produk.to_excel(buf2, index=False)
        st.download_button(
            "📥 Download Template Produk", buf2.getvalue(),
            "template_master_produk.xlsx", use_container_width=True)

        up_produk = st.file_uploader(
            "Upload Master Produk", type=["xlsx", "csv"], key="produk_upload")
        if up_produk:
            df = (pd.read_excel(up_produk) if up_produk.name.endswith("xlsx")
                  else pd.read_csv(up_produk))
            req = {"Kode_Produk", "Nama_Produk", "Kode_MC_Liquid",
                   "Grup_Cleaning", "Kg_per_CS", "Resting_Days", "Mixer_Kompatibel"}
            if req.issubset(df.columns):
                st.session_state.master_produk = df
                st.success(f"✅ {len(df)} produk berhasil diupload!")
            else:
                st.error(f"❌ Kolom harus: {req}")

        if not st.session_state.master_produk.empty:
            st.dataframe(st.session_state.master_produk,
                         use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════
# TAB 2 — INPUT PLANNING
# ═══════════════════════════════════════════════════════════════════════════
with tab2:
    st.header("Input Planning Filling")

    if st.session_state.master_produk.empty:
        st.warning("⚠️ Upload Master Produk dulu di tab Master Data.")
    else:
        # ── Week picker ──────────────────────────────────────────────────────
        st.subheader("📆 Pilih Minggu Filling")
        filling_week = st.date_input(
            "Pilih tanggal mana saja dalam minggu filling",
            value=datetime.today(),
            key="filling_week_picker")

        week_monday = filling_week - timedelta(days=filling_week.weekday())
        week_dates  = [week_monday + timedelta(days=i) for i in range(7)]
        week_labels = [
            f"{DAYS_ID[d.weekday()]} {d.strftime('%d/%m')}"
            for d in week_dates]

        shift_cols = []
        shift_meta = []
        for d, label in zip(week_dates, week_labels):
            for s in [1, 2, 3]:
                shift_cols.append(f"{label} S{s}")
                shift_meta.append((d.strftime("%Y-%m-%d"), s))

        st.caption(
            f"Minggu: **{week_monday.strftime('%d %b')} "
            f"— {week_dates[-1].strftime('%d %b %Y')}**")

        # ── Input kode produk ────────────────────────────────────────────────
        st.subheader("🔍 Pilih Produk yang Dijadwalkan")
        st.caption(
            "Ketik atau paste kode produk, pisahkan dengan koma atau baris baru. "
            "Produk yang sama boleh diinput lebih dari sekali (beda PO).")

        produk_df = st.session_state.master_produk
        all_kodes = list(produk_df["Kode_Produk"])

        raw_input = st.text_area(
            "Kode Produk",
            placeholder="Contoh: P001, P001, P002  (P001 dua kali = beda PO)",
            height=80,
            key=f"kode_input_{week_monday.strftime('%Y%m%d')}")

        col_btn1, col_btn2 = st.columns([3, 1])
        with col_btn2:
            tampilkan = st.button("🔍 Tampilkan Tabel", use_container_width=True)

        grid_key = f"grid_{week_monday.strftime('%Y%m%d')}"

        if tampilkan:
            input_kodes     = [
                k.strip()
                for k in re.split(r"[,\n\r\s]+", raw_input)
                if k.strip()]
            all_kodes_str   = [str(k) for k in all_kodes]
            input_kodes_str = [str(k) for k in input_kodes]
            not_found       = [
                k for k in input_kodes_str if k not in all_kodes_str]

            if not_found:
                st.warning(
                    f"⚠️ Tidak ditemukan di master: {', '.join(not_found)}")

            # FIX: preserve duplikasi — loop dari input_kodes_str bukan master
            nama_map_local = dict(zip(
                produk_df["Kode_Produk"].astype(str).str.strip(),
                produk_df["Nama_Produk"]))

            rows_kode = [k for k in input_kodes_str if k in all_kodes_str]
            rows_nama = [nama_map_local.get(k, k) for k in rows_kode]

            if rows_kode:
                init_data = {
                    "Urgent":      [False] * len(rows_kode),
                    "Kode_Produk": rows_kode,
                    "Nama_Produk": rows_nama,
                }
                for col in shift_cols:
                    init_data[col] = [None] * len(rows_kode)
                st.session_state[grid_key] = pd.DataFrame(init_data)
            else:
                st.warning("⚠️ Tidak ada kode produk yang valid.")

        # ── Template Excel & Upload ──────────────────────────────────────────
        if grid_key in st.session_state and not st.session_state[grid_key].empty:
            st.subheader("📋 Template Planning")
            template_df = st.session_state[grid_key].copy()

            wb = Workbook()
            ws = wb.active
            ws.title = "Planning"

            hdr_fill  = PatternFill("solid", fgColor="1F4E79")
            hdr_font  = Font(bold=True, color="FFFFFF")
            lock_fill = PatternFill("solid", fgColor="D9E1F2")
            center    = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

            headers = list(template_df.columns)
            for ci, h in enumerate(headers, 1):
                cell           = ws.cell(row=1, column=ci, value=h)
                cell.fill      = hdr_fill
                cell.font      = hdr_font
                cell.alignment = center

            for ri, (_, row) in enumerate(template_df.iterrows(), 2):
                for ci, (h, val) in enumerate(zip(headers, row), 1):
                    cell = ws.cell(
                        row=ri, column=ci,
                        value=val if val not in [None, False] else None)
                    cell.alignment = center
                    if h in ["Urgent", "Kode_Produk", "Nama_Produk"]:
                        cell.fill = lock_fill

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 20
            for i in range(3, len(headers)):
                ws.column_dimensions[get_column_letter(i + 1)].width = 12
            ws.row_dimensions[1].height = 30

            buf = io.BytesIO()
            wb.save(buf)

            st.download_button(
                "📥 Download Template Excel",
                buf.getvalue(),
                "template_planning.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

            st.caption(
                "Isi kolom CS di Excel, kolom **Urgent** isi `TRUE`/`FALSE`, "
                "lalu upload kembali.")

            up_filled = st.file_uploader(
                "📤 Upload Template yang Sudah Diisi",
                type=["xlsx"], key=f"up_{grid_key}")
            if up_filled:
                filled_df = pd.read_excel(up_filled)
                st.session_state[grid_key] = filled_df
                st.success("✅ Template berhasil diupload!")

            edited_df = st.session_state[grid_key]

        else:
            st.info("Masukkan kode produk lalu klik **Tampilkan Tabel**.")
            edited_df = pd.DataFrame()

        if st.button("💾 Simpan Planning", type="primary", use_container_width=True):
            if edited_df.empty:
                st.warning("⚠️ Masukkan kode produk dan isi tabel terlebih dahulu.")
            else:
                st.session_state.filling_plan = pd.DataFrame()
                rows = []
                for _, row in edited_df.iterrows():
                    kode   = str(row["Kode_Produk"]).strip()
                    nama   = str(row["Nama_Produk"]).strip()
                    urgent = (
                        "Urgent"
                        if str(row.get("Urgent", "")).strip().upper()
                           in ["TRUE", "1", "URGENT"]
                        else "Tidak Urgent")
                    for col, (date_str, shift_num) in zip(shift_cols, shift_meta):
                        val = row.get(col)
                        try:
                            val_float = (
                                float(val)
                                if pd.notna(val) and val is not None
                                else 0)
                        except (ValueError, TypeError):
                            val_float = 0
                        if val_float > 0:
                            rows.append({
                                "Kode_Produk":     kode,
                                "Nama_Produk":     nama,
                                "Target_CS":       val_float,
                                "Tanggal_Filling": date_str,
                                "Shift_Filling":   shift_num,
                                "Urgent":          urgent
                            })

                if rows:
                    st.session_state.filling_plan = pd.DataFrame(rows)
                    st.success(f"✅ {len(rows)} item planning tersimpan!")
                else:
                    st.warning("⚠️ Tidak ada data yang diisi.")

        if not st.session_state.filling_plan.empty:
            with st.expander("📋 Lihat Planning Tersimpan"):
                st.dataframe(st.session_state.filling_plan,
                             use_container_width=True, hide_index=True)
                st.caption(
                    f"Total: {len(st.session_state.filling_plan)} item")

# ═══════════════════════════════════════════════════════════════════════════
# TAB 3 — JADWAL MIXING
# ═══════════════════════════════════════════════════════════════════════════
with tab3:
    st.header("Jadwal Mixing Otomatis")

    ready = (not st.session_state.master_mixer.empty and
             not st.session_state.master_produk.empty and
             not st.session_state.filling_plan.empty)

    if not ready:
        st.warning(
            "⚠️ Lengkapi Master Mixer, Master Produk, "
            "dan Input Planning terlebih dahulu.")
    else:
        st.subheader("📆 Range Jadwal Mixing")

        filling_plan_df = st.session_state.filling_plan
        min_fill_date   = pd.to_datetime(
            filling_plan_df["Tanggal_Filling"]).min()
        days_to_friday  = (min_fill_date.weekday() - 4) % 7
        default_start   = (
            min_fill_date - timedelta(days=days_to_friday + 7)).date()
        default_end     = default_start + timedelta(days=9)

        c1, c2 = st.columns(2)
        with c1:
            mix_start = st.date_input(
                "Dari tanggal (Jumat)", value=default_start)
        with c2:
            mix_end = st.date_input("Sampai tanggal", value=default_end)

        date_range = []
        d = mix_start
        while d <= mix_end:
            date_range.append(d.strftime("%Y-%m-%d"))
            d += timedelta(days=1)
        st.caption(
            f"Range mixing: **{mix_start.strftime('%d %b')} "
            f"— {mix_end.strftime('%d %b %Y')}** ({len(date_range)} hari)")

        if st.button("⚡ Generate Jadwal Mixing", type="primary",
                     use_container_width=True):
            with st.spinner("Menjadwalkan mixing..."):
                result = generate_mixing_schedule(
                    st.session_state.master_mixer,
                    st.session_state.master_produk,
                    st.session_state.filling_plan,
                    date_range=date_range)
            st.session_state.schedule_result = result

        if "schedule_result" in st.session_state:
            result      = st.session_state.schedule_result
            schedule_df = result["schedule"]
            warnings    = result["warnings"]
            shifted     = result["shifted"]
            unscheduled = result["unscheduled"]

            if warnings:
                for w in warnings:
                    st.warning(w)

            if shifted:
                st.subheader("🔀 Jadwal Filling Digeser (Tidak Urgent)")
                st.dataframe(
                    pd.DataFrame(shifted).style.map(
                        lambda _: "background-color: #fff3cd"),
                    use_container_width=True, hide_index=True)

            if unscheduled:
                st.subheader("❌ Tidak Bisa Dijadwalkan")
                for u in unscheduled:
                    st.error(u)

            if not schedule_df.empty:
                st.subheader("📅 Jadwal Mixing")

                pivot_df, meta = build_pivot(
                    schedule_df,
                    st.session_state.master_mixer,
                    st.session_state.master_produk,
                    date_range)

                if not pivot_df.empty:
                    display_df  = pivot_df.copy()
                    rename_map  = {
                        c: c.replace("\n", " ")
                        for c in display_df.columns}
                    display_df  = display_df.rename(columns=rename_map)
                    col_display = [
                        c.replace("\n", " ")
                        for c in meta["col_labels"]]

                    def style_pivot(row):
                        styles = [""] * len(row)
                        jid    = row.get("Job_ID", 0)
                        cols   = list(row.index)
                        for i, label in enumerate(col_display):
                            d_key, s = meta["col_keys"][i]
                            if label not in cols:
                                continue
                            pos = cols.index(label)
                            if (jid, d_key, s) in meta["cleaning_cells"]:
                                styles[pos] = "background-color: #BDD7EE"
                            elif (jid, d_key, s) in meta["resting_cells"]:
                                styles[pos] = "background-color: #FFE699"
                        return styles

                    st.dataframe(
                        display_df.style.apply(style_pivot, axis=1),
                        use_container_width=True, hide_index=True)
                    st.caption("🔵 Biru = Cleaning | 🟡 Kuning = Resting")

                    excel_data = pivot_to_excel(
                        pivot_df, meta,
                        st.session_state.master_mixer,
                        filling_plan=st.session_state.filling_plan,
                        master_produk=st.session_state.master_produk)
                    st.download_button(
                        "📥 Download Excel",
                        excel_data,
                        "jadwal_mixing.xlsx",
                        mime="application/vnd.openxmlformats-officedocument"
                             ".spreadsheetml.sheet",
                        use_container_width=True)

                with st.expander("📋 Detail Jadwal (Raw)"):
                    st.dataframe(
                        schedule_df.drop(
                            columns=["Cleaning"], errors="ignore"),
                        use_container_width=True, hide_index=True)
                    st.write("**Debug:**")
                    st.write("Mixer di master:",
                             list(st.session_state.master_mixer["Mixer"]))
                    st.write("Sample Mixer_Kompatibel:",
                             list(st.session_state.master_produk[
                                 "Mixer_Kompatibel"].head(3)))
                    st.write("Sample Kode di filling plan:",
                             list(st.session_state.filling_plan[
                                 "Kode_Produk"].head(3)))
                    st.write("Sample Kode di master produk:",
                             list(st.session_state.master_produk[
                                 "Kode_Produk"].astype(str).head(3)))

                with st.expander("🔍 Debug: Pivot Rows"):
                    if not pivot_df.empty:
                        st.write("Baris di pivot:",
                                 pivot_df[["Job_ID", "Mixer",
                                           "Kode_Produk",
                                           "Nama_Produk"]].to_dict("records"))
                    else:
                        st.write("Pivot kosong!")
                    st.write("Unscheduled:", result["unscheduled"])
