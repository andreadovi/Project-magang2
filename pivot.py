# pivot.py
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

DAYS_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]

# ─── Style constants ──────────────────────────────────────────────────────────
FILL_HEADER_DARK  = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER_LIGHT = PatternFill("solid", fgColor="BDD7EE")
FILL_CLEANING     = PatternFill("solid", fgColor="BDD7EE")
FILL_RESTING      = PatternFill("solid", fgColor="FFE699")
FILL_MIXING       = PatternFill("solid", fgColor="E2EFDA")
FONT_WHITE_BOLD   = Font(bold=True, color="FFFFFF")
FONT_BOLD         = Font(bold=True)
ALIGN_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ─── Build pivot DataFrame ────────────────────────────────────────────────────

def build_pivot(schedule_df, master_mixer_df, master_produk_df, date_range):
    """
    Baris  = (Mixer, Kode_Produk, Nama_Produk)
    Kolom  = setiap (tanggal, shift) dalam date_range

    Returns:
        pivot_df : DataFrame siap tampil
        meta     : dict { col_keys, col_labels, cleaning_cells, resting_cells, date_range }
    """
    if schedule_df.empty:
        return pd.DataFrame(), {}

    mixer_df = master_mixer_df.copy()
    mixer_df["Mixer"] = mixer_df["Mixer"].astype(str).str.strip()
    mixer_order = list(mixer_df["Mixer"])

    # ── Bangun kolom ──────────────────────────────────────────────────────────
    col_keys   = []   # list of (date_str, shift_int)
    col_labels = []   # label tampilan dengan \n

    for d_str in date_range:
        d = datetime.strptime(d_str, "%Y-%m-%d")
        day_lbl = f"{DAYS_ID[d.weekday()]}\n{d.strftime('%d/%m')}"
        for s in [1, 2, 3]:
            col_keys.append((d_str, s))
            col_labels.append(f"{day_lbl}\nS{s}")

    # ── Normalisasi schedule ──────────────────────────────────────────────────
    sdf = schedule_df.copy()
    sdf["Tanggal_Mixing"] = sdf["Tanggal_Mixing"].astype(str).str.strip()
    sdf["Shift_Mixing"]   = pd.to_numeric(sdf["Shift_Mixing"], errors="coerce").fillna(1).astype(int)
    sdf["Mixer"]          = sdf["Mixer"].astype(str).str.strip()
    sdf["Kode_Produk"]    = sdf["Kode_Produk"].astype(str).str.strip()
    sdf["Kg_Mixing"]      = pd.to_numeric(sdf["Kg_Mixing"], errors="coerce").fillna(0)

    # ── Kumpulkan kombinasi (mixer, kode_produk) unik ─────────────────────────
    combos = (
        sdf[["Mixer", "Kode_Produk", "Nama_Produk", "Kode_MC_Liquid",
             "Grup_Cleaning", "Resting_Days", "Tanggal_Filling", "Shift_Filling"]]
        .drop_duplicates(subset=["Mixer", "Kode_Produk"])
        .copy()
    )
    combos["_mixer_order"] = combos["Mixer"].apply(
        lambda m: mixer_order.index(m) if m in mixer_order else 999
    )
    combos = combos.sort_values(["_mixer_order", "Kode_Produk"]).drop(columns=["_mixer_order"])

    cleaning_cells = set()   # (mixer, kode, date_str, shift)
    resting_cells  = set()

    rows = []
    for _, combo in combos.iterrows():
        mixer      = combo["Mixer"]
        kode       = combo["Kode_Produk"]
        nama       = combo["Nama_Produk"]
        kode_mc    = combo.get("Kode_MC_Liquid", "")
        grup       = combo["Grup_Cleaning"]
        rest_days  = int(float(combo.get("Resting_Days", 0)))
        fill_date  = pd.to_datetime(combo["Tanggal_Filling"])

        row_data = {
            "Mixer":          mixer,
            "Kode_Produk":    kode,
            "Kode_MC_Liquid": kode_mc,
            "Nama_Produk":    nama,
            "Grup_Cleaning":  grup,
        }

        mask = (sdf["Mixer"] == mixer) & (sdf["Kode_Produk"] == kode)
        sub  = sdf[mask]

        for (d_str, s) in col_keys:
            cell_mask = (sub["Tanggal_Mixing"] == d_str) & (sub["Shift_Mixing"] == s)
            cell_data = sub[cell_mask]

            if not cell_data.empty:
                kg_val         = cell_data["Kg_Mixing"].sum()
                cleaning_flag  = bool(cell_data["Cleaning"].any())
                if cleaning_flag:
                    row_data[(d_str, s)] = f"🔵 {round(kg_val, 1)} kg"
                    cleaning_cells.add((mixer, kode, d_str, s))
                else:
                    row_data[(d_str, s)] = f"{round(kg_val, 1)} kg"
            else:
                # Cek resting period
                d_dt       = datetime.strptime(d_str, "%Y-%m-%d")
                in_resting = False
                if rest_days > 0 and not sub.empty:
                    last_mix   = pd.to_datetime(sub["Tanggal_Mixing"]).max()
                    rest_start = last_mix + timedelta(days=1)
                    if rest_start <= d_dt <= fill_date:
                        in_resting = True
                if in_resting:
                    row_data[(d_str, s)] = "💤 Resting"
                    resting_cells.add((mixer, kode, d_str, s))
                else:
                    row_data[(d_str, s)] = ""

        rows.append(row_data)

    # ── Build DataFrame ───────────────────────────────────────────────────────
    rename_map = {ck: cl for ck, cl in zip(col_keys, col_labels)}
    pivot_df   = pd.DataFrame(rows).rename(columns=rename_map)

    meta = {
        "col_keys":       col_keys,
        "col_labels":     col_labels,
        "cleaning_cells": cleaning_cells,
        "resting_cells":  resting_cells,
        "date_range":     date_range,
    }
    return pivot_df, meta


# ─── Export ke Excel ──────────────────────────────────────────────────────────

def pivot_to_excel(pivot_df, meta, master_mixer_df):
    """Export pivot ke Excel dengan formatting lengkap."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Jadwal Mixing"

    mixer_df = master_mixer_df.copy()
    mixer_df["Mixer"] = mixer_df["Mixer"].astype(str).str.strip()

    col_keys        = meta["col_keys"]
    date_range      = meta["date_range"]
    cleaning_cells  = meta["cleaning_cells"]
    resting_cells   = meta["resting_cells"]

    FIXED_COLS = ["Mixer", "Kode_Produk", "Kode_MC_Liquid", "Nama_Produk", "Grup_Cleaning"]
    n_fixed    = len(FIXED_COLS)

    # ── Row 1: kolom tetap (merge 2 baris) + header tanggal (merge 3 kolom) ───
    for ci, h in enumerate(FIXED_COLS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = FILL_HEADER_DARK
        c.font      = FONT_WHITE_BOLD
        c.alignment = ALIGN_CENTER
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

    col_cursor = n_fixed + 1
    for d_str in date_range:
        d     = datetime.strptime(d_str, "%Y-%m-%d")
        label = f"{DAYS_ID[d.weekday()]} {d.strftime('%d/%m/%Y')}"
        c = ws.cell(row=1, column=col_cursor, value=label)
        c.fill      = FILL_HEADER_DARK
        c.font      = FONT_WHITE_BOLD
        c.alignment = ALIGN_CENTER
        ws.merge_cells(
            start_row=1, start_column=col_cursor,
            end_row=1,   end_column=col_cursor + 2
        )
        col_cursor += 3

    # ── Row 2: shift header ───────────────────────────────────────────────────
    for ci, (d_str, s) in enumerate(col_keys):
        c = ws.cell(row=2, column=n_fixed + 1 + ci, value=f"S{s}")
        c.fill      = FILL_HEADER_LIGHT
        c.font      = FONT_BOLD
        c.alignment = ALIGN_CENTER

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(pivot_df.iterrows(), 3):
        mixer = str(row.get("Mixer", ""))
        kode  = str(row.get("Kode_Produk", ""))

        for ci, col in enumerate(FIXED_COLS, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            c.alignment = ALIGN_LEFT
            c.border    = THIN_BORDER

        for ci, (col_label, (d_str, s)) in enumerate(
            zip(meta["col_labels"], col_keys), n_fixed + 1
        ):
            val = row.get(col_label, "")
            c   = ws.cell(row=ri, column=ci, value=val)
            c.alignment = ALIGN_CENTER
            c.border    = THIN_BORDER

            if (mixer, kode, d_str, s) in cleaning_cells:
                c.fill = FILL_CLEANING
            elif (mixer, kode, d_str, s) in resting_cells:
                c.fill = FILL_RESTING
            elif val and val != "":
                c.fill = FILL_MIXING

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10   # Mixer
    ws.column_dimensions["B"].width = 14   # Kode_Produk
    ws.column_dimensions["C"].width = 14   # Kode_MC_Liquid
    ws.column_dimensions["D"].width = 30   # Nama_Produk
    ws.column_dimensions["E"].width = 12   # Grup_Cleaning
    for ci in range(n_fixed + 1, n_fixed + len(col_keys) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "F3"

    # ── Sheet Keterangan ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Keterangan")
    legends = [
        ("Warna",           "Arti"),
        ("Hijau muda",      "Ada jadwal mixing (kg)"),
        ("Biru muda (🔵)", "Ada cleaning sebelum mixing di slot ini"),
        ("Kuning (💤)",    "Periode resting (menunggu filling)"),
        ("Kosong",          "Tidak ada aktivitas"),
    ]
    for ri, (a, b) in enumerate(legends, 1):
        ws2.cell(row=ri, column=1, value=a).font = FONT_BOLD
        ws2.cell(row=ri, column=2, value=b)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
